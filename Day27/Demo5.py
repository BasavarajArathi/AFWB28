#READING DATA FROM EXCEL
import openpyxl

wb=openpyxl.load_workbook("F:/B28/Book1.xlsx")

sheet1=wb['Sheet1']
rc=sheet1.max_row
print("Row Count",rc)
cc=sheet1.max_column
print("Cplumn count",cc)
for i in range(1,rc+1):
    for j in range(1,cc+1):
        v=sheet1.cell(i,j).value
        print(v,end=" ")
    print()
wb.close()

#A1 B1  C1
#A2 B2  C2
#A3 B3  C3